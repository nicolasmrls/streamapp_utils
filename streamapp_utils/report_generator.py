from io import BytesIO
from pathlib import Path
from datetime import date
from pandas import DataFrame
from typing import Any
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, NamedStyle, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from streamlit import secrets, toast
from zipfile import ZipFile, ZipInfo


class ReportGenerator:
    date = date.today().isoformat()
    # stylish
    totals_style = NamedStyle(name="totals_style")
    totals_style.font = Font(name='Arial', size=13, bold=True)
    totals_style.border = Border(
        top=Side(border_style='thick', color='FF000000'),
        bottom=Side(border_style=None, color='FF000000')
    )
    col = get_column_letter

    @staticmethod
    def __sheet_names(name: str):
        try:
            assert name.find(' ') == -1, 'Sheet names cannot have spaces'
        except AssertionError as e:
            return e
        return

    @staticmethod
    def __reports(workbook: Any, dfs: dict, file_name: str,
                  base_file: str = '_',
                  xslx_style: str = 'TableStyleMedium3'):

        for name, df in dfs.items():
            val = ReportGenerator.__sheet_names(name)
            if val:
                toast(val, icon='⛔')
                return [file_name+'.txt', b'']
            for i in df.columns:
                try:
                    df[i] = df[i].dt.date
                except Exception:
                    pass
            (max_row, max_col) = df.shape
            worksheet = workbook[name]

            for i in range(1, max_col+1):
                worksheet.column_dimensions[get_column_letter(i)].width = 23

            tab = Table(
                displayName=name,
                ref=f"A1:{get_column_letter(max_col)}{max_row+1}"
            )
            tab.tableStyleInfo = TableStyleInfo(name=xslx_style)
            worksheet._tables.add(tab)

            header = base_file == '_'
            base_df = dataframe_to_rows(df, index=False, header=header)
            for c, r in enumerate(base_df, int(not header)):
                for i, j in enumerate(r, 1):
                    worksheet.cell(c+1, i, j)

        with BytesIO() as buffer:
            workbook.save(buffer)
            file = buffer.getvalue()
        workbook.close()

        return [file_name+'.xlsx', file]

    @staticmethod
    def from_template_xlsx(file_name: str, dfs: dict,
                           sub_folder: str = '', base_file: str = '_'):
        """
        Return an .xlsx file
        dfs is a dict of names and Dataframes per sheets
        ex: {'Hoja 1': Dataframe1, 'ingresos': ingresos_dataframe}
        kargs: variables to define a summary

        """
        try:
            workbook = load_workbook(
                Path(
                    secrets.utils_files,
                    sub_folder, base_file + '.xlsx'
                ).as_posix()
            )
            result = ReportGenerator.__reports(
                workbook=workbook,
                dfs=dfs,
                file_name=file_name,
                base_file=base_file
            )
        except Exception:
            result = ReportGenerator.xlsx(
                dfs=dfs,
                file_name=file_name
            )

        return result

    @staticmethod
    def xlsx(dfs: dict[str, DataFrame], file_name: str):
        """
        Return an .xlsx file
        dfs is a dict of names and Dataframes per sheets
        ex: {'Seeht_1': Dataframe1, 'ingresos': ingresos_dataframe}
        kargs: variables to define a summary

        """
        workbook = Workbook()
        for n, i in enumerate(dfs.keys()):
            val = ReportGenerator.__sheet_names(i)
            if val:
                toast(val, icon='⛔')
                return [file_name+'.txt', b'']
            if n == 0:
                workbook.active.__setattr__('title', i)
            else:
                workbook.create_sheet(title=i)
        result = ReportGenerator.__reports(
            workbook=workbook,
            dfs=dfs,
            file_name=file_name
        )

        return result


class InMemoryZip(object):
    zip_file = BytesIO()

    @classmethod
    def create_zip(cls, reports: list):
        """
        returns: zip archive
        """
        with ZipFile(cls.zip_file, 'w') as zip_archive:
            for i in reports:
                zip_archive.writestr(
                    zinfo_or_arcname=ZipInfo(i[0]),
                    data=i[1]
                )
        return cls.zip_file
